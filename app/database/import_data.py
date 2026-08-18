from decimal import Decimal, InvalidOperation
from pathlib import Path
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.database.database import engine
from app.database.models import Invoice


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_DATA_DIR = PROJECT_ROOT / "data" / "sample"

REQUIRED_COLUMNS = [
    "Customer ID",
    "Customer Name",
    "Invoice No",
    "Invoice Date",
    "Due Date",
    "Invoice Amount",
    "Amount Paid",
    "Outstanding Amount",
    "Aging Days",
    "Aging Bucket",
    "Currency",
    "Payment Status",
]


def find_excel_file() -> Path:
    """Find the single Excel source file in the sample directory."""

    excel_files = list(SAMPLE_DATA_DIR.glob("*.xlsx"))

    if not excel_files:
        raise FileNotFoundError(
            f"No Excel file found in:\n{SAMPLE_DATA_DIR}"
        )

    if len(excel_files) > 1:
        filenames = "\n".join(
            f"  - {file.name}" for file in excel_files
        )

        raise ValueError(
            "Multiple Excel files were found.\n"
            "Keep only the intended source dataset.\n\n"
            f"{filenames}"
        )

    return excel_files[0]


def parse_date(value, field_name: str, row_number: int) -> date:
    """Convert an Excel date value into a Python date."""

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    raise ValueError(
        f"Invalid {field_name} at Excel row {row_number}: {value!r}"
    )


def parse_decimal(value, field_name: str, row_number: int) -> Decimal:
    """Convert a financial value into Decimal with two decimal places."""

    if value is None:
        raise ValueError(
            f"Missing {field_name} at Excel row {row_number}."
        )

    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"Invalid {field_name} at Excel row {row_number}: {value!r}"
        ) from exc

    return decimal_value.quantize(Decimal("0.01"))


def parse_integer(value, field_name: str, row_number: int) -> int:
    """Convert a value into an integer."""

    if value is None:
        raise ValueError(
            f"Missing {field_name} at Excel row {row_number}."
        )

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Invalid {field_name} at Excel row {row_number}: {value!r}"
        ) from exc


def validate_headers(headers: list) -> None:
    """Verify that all required columns are present."""

    actual_columns = {
        str(header).strip()
        for header in headers
        if header is not None
    }

    missing_columns = set(REQUIRED_COLUMNS) - actual_columns

    if missing_columns:
        raise ValueError(
            f"Missing required columns: {sorted(missing_columns)}"
        )


def import_data() -> None:
    """Import validated Excel records into SQLite."""

    print("=" * 70)
    print("FINANCE DATA IMPORT")
    print("=" * 70)

    excel_file = find_excel_file()

    print(f"\nSource file:")
    print(excel_file)

    workbook = load_workbook(
        excel_file,
        read_only=True,
        data_only=True,
    )

    try:
        if "Customer Aging Report" not in workbook.sheetnames:
            raise ValueError(
                "The 'Customer Aging Report' sheet was not found."
            )

        worksheet = workbook["Customer Aging Report"]

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

        headers = list(header_row)
        validate_headers(headers)

        column_indexes = {
            column: headers.index(column)
            for column in REQUIRED_COLUMNS
        }

        records = []
        invoice_numbers = set()

        print("\nReading and validating records...")

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if not any(value is not None for value in row):
                continue

            customer_id = row[column_indexes["Customer ID"]]
            customer_name = row[column_indexes["Customer Name"]]
            invoice_no = row[column_indexes["Invoice No"]]

            if not customer_id:
                raise ValueError(
                    f"Missing Customer ID at Excel row {row_number}."
                )

            if not customer_name:
                raise ValueError(
                    f"Missing Customer Name at Excel row {row_number}."
                )

            if not invoice_no:
                raise ValueError(
                    f"Missing Invoice No at Excel row {row_number}."
                )

            invoice_no = str(invoice_no).strip()

            if invoice_no in invoice_numbers:
                raise ValueError(
                    f"Duplicate invoice number found at Excel row "
                    f"{row_number}: {invoice_no}"
                )

            invoice_numbers.add(invoice_no)

            invoice_date = parse_date(
                row[column_indexes["Invoice Date"]],
                "Invoice Date",
                row_number,
            )

            due_date = parse_date(
                row[column_indexes["Due Date"]],
                "Due Date",
                row_number,
            )

            invoice_amount = parse_decimal(
                row[column_indexes["Invoice Amount"]],
                "Invoice Amount",
                row_number,
            )

            amount_paid = parse_decimal(
                row[column_indexes["Amount Paid"]],
                "Amount Paid",
                row_number,
            )

            outstanding_amount = parse_decimal(
                row[column_indexes["Outstanding Amount"]],
                "Outstanding Amount",
                row_number,
            )

            aging_days = parse_integer(
                row[column_indexes["Aging Days"]],
                "Aging Days",
                row_number,
            )

            aging_bucket = str(
                row[column_indexes["Aging Bucket"]]
            ).strip()

            currency = str(
                row[column_indexes["Currency"]]
            ).strip()

            payment_status = str(
                row[column_indexes["Payment Status"]]
            ).strip()

            if invoice_amount < 0:
                raise ValueError(
                    f"Negative Invoice Amount at Excel row {row_number}."
                )

            if amount_paid < 0:
                raise ValueError(
                    f"Negative Amount Paid at Excel row {row_number}."
                )

            if outstanding_amount < 0:
                raise ValueError(
                    f"Negative Outstanding Amount at Excel row "
                    f"{row_number}."
                )

            calculated_outstanding = (
                invoice_amount - amount_paid
            ).quantize(Decimal("0.01"))

            if calculated_outstanding != outstanding_amount:
                raise ValueError(
                    f"Outstanding amount mismatch at Excel row "
                    f"{row_number}. "
                    f"Expected {calculated_outstanding}, "
                    f"found {outstanding_amount}."
                )

            record = Invoice(
                customer_id=str(customer_id).strip(),
                customer_name=str(customer_name).strip(),
                invoice_no=invoice_no,
                invoice_date=invoice_date,
                due_date=due_date,
                invoice_amount=invoice_amount,
                amount_paid=amount_paid,
                outstanding_amount=outstanding_amount,
                aging_days=aging_days,
                aging_bucket=aging_bucket,
                currency=currency,
                payment_status=payment_status,
            )

            records.append(record)

        print(f"Validated records: {len(records)}")

        if not records:
            raise ValueError("No records were found in the Excel file.")

        with Session(engine) as session:
            existing_invoice_numbers = set(
                session.scalars(
                    select(Invoice.invoice_no).where(
                        Invoice.invoice_no.in_(invoice_numbers)
                    )
                ).all()
            )

            if existing_invoice_numbers:
                sample = sorted(existing_invoice_numbers)[:10]

                raise ValueError(
                    "Import stopped because some invoice numbers "
                    "already exist in the database.\n"
                    f"Existing invoices: {sample}"
                )

            session.add_all(records)
            session.commit()

        print(f"Imported records: {len(records)}")
        print("\n" + "=" * 70)
        print("IMPORT SUCCESSFUL")
        print("=" * 70)

    except (SQLAlchemyError, ValueError, TypeError) as exc:
        print("\nIMPORT FAILED")
        print(f"Reason: {exc}")
        raise

    finally:
        workbook.close()


if __name__ == "__main__":
    import_data()