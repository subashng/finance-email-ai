from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_DATA_DIR = PROJECT_ROOT / "data" / "sample"

REQUIRED_SHEETS = {
    "Customer Aging Report",
    "Customer Summary",
    "Data Dictionary",
}

REQUIRED_COLUMNS = {
    "Customer ID",
    "Customer Name",
    "Invoice No",
    "Invoice Date",
    "Due Date",
    "Invoice Amount",
    "Amount Paid",
    "Outstanding Amount",
    "Aging Days",
    "Aging Bucket",
    "Currency",
    "Payment Status",
}


def find_excel_file() -> Path:
    """Find the single Excel source file in the sample data directory."""

    excel_files = list(SAMPLE_DATA_DIR.glob("*.xlsx"))

    if not excel_files:
        raise FileNotFoundError(
            f"No Excel (.xlsx) file was found in:\n{SAMPLE_DATA_DIR}"
        )

    if len(excel_files) > 1:
        filenames = "\n".join(f"  - {file.name}" for file in excel_files)

        raise ValueError(
            "Multiple Excel files were found in the sample data directory.\n"
            "Keep only the intended source dataset.\n\n"
            f"{filenames}"
        )

    return excel_files[0]


def validate_workbook(excel_file: Path) -> None:
    """Validate the structure and basic integrity of the Excel dataset."""

    print("=" * 70)
    print("FINANCE DATA SOURCE VALIDATION")
    print("=" * 70)

    print("\nFile:")
    print(excel_file)

    if not excel_file.exists():
        raise FileNotFoundError(
            f"Excel file was not found:\n{excel_file}"
        )

    print("PASS: Excel file exists.")

    workbook = load_workbook(
        excel_file,
        read_only=True,
        data_only=True,
    )

    try:
        sheet_names = set(workbook.sheetnames)

        print("\nSheets found:")
        for sheet_name in workbook.sheetnames:
            print(f"  - {sheet_name}")

        missing_sheets = REQUIRED_SHEETS - sheet_names

        if missing_sheets:
            raise ValueError(
                f"Missing required sheets: {sorted(missing_sheets)}"
            )

        print("PASS: Required sheets are present.")

        worksheet = workbook["Customer Aging Report"]

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

        headers = list(header_row)

        actual_columns = {
            str(header).strip()
            for header in headers
            if header is not None
        }

        print("\nColumns found:")
        for column in headers:
            print(f"  - {column}")

        missing_columns = REQUIRED_COLUMNS - actual_columns

        if missing_columns:
            raise ValueError(
                f"Missing required columns: {sorted(missing_columns)}"
            )

        print("PASS: Required columns are present.")

        customer_id_index = headers.index("Customer ID")
        invoice_index = headers.index("Invoice No")

        customer_ids = set()
        invoice_numbers = set()
        data_rows = 0

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            if not any(value is not None for value in row):
                continue

            data_rows += 1

            customer_id = row[customer_id_index]
            invoice_number = row[invoice_index]

            if customer_id is not None:
                customer_ids.add(str(customer_id).strip())

            if invoice_number is not None:
                invoice_numbers.add(str(invoice_number).strip())

        print(f"\nInvoice/data line items found: {data_rows}")
        print(f"Unique customers found: {len(customer_ids)}")
        print(f"Unique invoices found: {len(invoice_numbers)}")

        if len(customer_ids) < 200:
            raise ValueError(
                f"Expected at least 200 unique customers, "
                f"but found {len(customer_ids)}."
            )

        print("PASS: At least 200 unique customers are present.")

        if data_rows < 200:
            raise ValueError(
                f"Expected at least 200 data rows, "
                f"but found {data_rows}."
            )

        print("PASS: At least 200 invoice/data line items are present.")

        if len(invoice_numbers) != data_rows:
            raise ValueError(
                "Invoice numbers are not unique across the dataset."
            )

        print("PASS: Invoice numbers are unique.")

        print("\n" + "=" * 70)
        print("VALIDATION SUCCESSFUL")
        print("=" * 70)
        print(
            "The Excel source data is structurally ready "
            "for database import."
        )

    finally:
        workbook.close()


def main() -> None:
    excel_file = find_excel_file()
    validate_workbook(excel_file)


if __name__ == "__main__":
    main()